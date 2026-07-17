from datetime import datetime
from pathlib import Path

from PySide6.QtGui import QTextDocument
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import QFileDialog, QMessageBox

from modules.equipo_utils import icono_tipo_archivo


def exportar_equipos_excel(db, parent) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        QMessageBox.critical(
            parent, "Excel", "Instale openpyxl: pip install openpyxl"
        )
        return

    filas = db.obtener_equipos()
    if not filas:
        QMessageBox.information(parent, "Excel", "No hay equipos para exportar.")
        return

    nombre = f"equipos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta, _ = QFileDialog.getSaveFileName(
        parent,
        "Exportar Equipos a Excel",
        str(Path.home() / nombre),
        "Excel (*.xlsx)",
    )
    if not ruta:
        return
    if not ruta.lower().endswith(".xlsx"):
        ruta += ".xlsx"

    encabezados = [
        "Código", "Nombre", "Tipo", "Marca", "Modelo", "Año", "Serie",
        "Fabricante", "Motor", "Serie motor", "Potencia", "Proyecto",
        "Ubicación", "Horómetro", "Estado", "Observaciones",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipos"

    for col, tit in enumerate(encabezados, start=1):
        c = ws.cell(row=1, column=col, value=tit)
        c.font = Font(bold=True)

    for ri, row in enumerate(filas, start=2):
        ws.cell(row=ri, column=1, value=row["codigo"])
        ws.cell(row=ri, column=2, value=row["nombre"])
        ws.cell(row=ri, column=3, value=row["tipo"])
        ws.cell(row=ri, column=4, value=row["marca"])
        ws.cell(row=ri, column=5, value=row["modelo"])
        ws.cell(row=ri, column=6, value=row["anio"])
        ws.cell(row=ri, column=7, value=row["serie"])
        ws.cell(row=ri, column=8, value=row["fabricante"])
        ws.cell(row=ri, column=9, value=row["motor"])
        ws.cell(row=ri, column=10, value=row["serie_motor"])
        ws.cell(row=ri, column=11, value=row["potencia"])
        ws.cell(row=ri, column=12, value=row["proyecto"])
        ws.cell(row=ri, column=13, value=row["ubicacion"])
        ws.cell(row=ri, column=14, value=db.obtener_horometro_actual(row["id"]))
        ws.cell(row=ri, column=15, value=row["estado"])
        ws.cell(row=ri, column=16, value=row["observaciones"])

    wb.save(ruta)
    QMessageBox.information(parent, "Excel", f"Exportado:\n{ruta}")


def exportar_ficha_pdf(db, equipo_id: int, parent) -> None:
    row = db.obtener_equipo_por_id(equipo_id)
    if row is None:
        return

    horo = db.obtener_horometro_actual(equipo_id)
    fotos = db.obtener_fotografias_equipo(equipo_id)
    ruta_foto = ""
    for f in fotos:
        if f["categoria"] == "Equipo" and Path(f["ruta"]).is_file():
            ruta_foto = f["ruta"]
            break
    if not ruta_foto and fotos:
        ruta_foto = fotos[0]["ruta"]

    img_html = ""
    if ruta_foto and Path(ruta_foto).is_file():
        p = Path(ruta_foto).resolve().as_posix()
        img_html = f'<p><img src="file:///{p}" width="280"/></p>'

    docs = []
    for titulo, clave in (
        ("Manual de Operación", "ruta_manual"),
        ("Manual de Partes", "ruta_manual_partes"),
        ("Plano Hidráulico", "ruta_plano_hidraulico"),
        ("Plano Eléctrico", "ruta_plano_electrico"),
    ):
        r = row[clave] if clave in row.keys() else ""
        if r:
            docs.append(f"<li>{titulo}: {icono_tipo_archivo(r)} {Path(r).name}</li>")

    docs_html = (
        "<ul>" + "".join(docs) + "</ul>"
        if docs else "<p>Sin documentos adjuntos.</p>"
    )

    nombre = (
        f"ficha_{row['codigo']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    )
    ruta, _ = QFileDialog.getSaveFileName(
        parent,
        "Exportar ficha técnica PDF",
        str(Path.home() / nombre),
        "PDF (*.pdf)",
    )
    if not ruta:
        return
    if not ruta.lower().endswith(".pdf"):
        ruta += ".pdf"

    html = f"""
    <html><head><meta charset="utf-8"/>
    <style>
      body {{ font-family: Segoe UI, Arial; font-size: 10pt; }}
      h1 {{ color: #2563EB; font-size: 16pt; }}
      h2 {{ color: #334155; font-size: 12pt; margin-top: 16px; }}
      table {{ border-collapse: collapse; width: 100%; }}
      td {{ padding: 4px 8px; vertical-align: top; }}
      .lbl {{ font-weight: bold; width: 160px; }}
    </style></head><body>
    <h1>Ficha técnica — SGM Piedra Vial</h1>
    <h2>{row['codigo']} — {row['nombre']}</h2>
    {img_html}
    <h2>Datos generales</h2>
    <table>
      <tr><td class="lbl">Tipo</td><td>{row['tipo'] or ''}</td></tr>
      <tr><td class="lbl">Marca / Modelo</td><td>{row['marca'] or ''} {row['modelo'] or ''}</td></tr>
      <tr><td class="lbl">Proyecto</td><td>{row['proyecto'] or ''}</td></tr>
      <tr><td class="lbl">Ubicación</td><td>{row['ubicacion'] or ''}</td></tr>
      <tr><td class="lbl">Estado</td><td>{row['estado'] or ''}</td></tr>
      <tr><td class="lbl">Horómetro actual</td><td>{horo}</td></tr>
    </table>
    <h2>Motor</h2>
    <table>
      <tr><td class="lbl">Motor</td><td>{row['motor'] or ''}</td></tr>
      <tr><td class="lbl">Serie motor</td><td>{row['serie_motor'] or ''}</td></tr>
      <tr><td class="lbl">Potencia</td><td>{row['potencia'] or ''}</td></tr>
    </table>
    <h2>Documentos</h2>
    {docs_html}
    </body></html>
    """

    doc = QTextDocument()
    doc.setHtml(html)
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(ruta)
    doc.print_(printer)
    QMessageBox.information(parent, "PDF", f"Ficha exportada:\n{ruta}")
