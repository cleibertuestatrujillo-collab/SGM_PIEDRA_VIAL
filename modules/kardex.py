from datetime import datetime
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtGui import QTextDocument
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QComboBox,
    QMessageBox,
    QFileDialog,
    QDialog,
    QFrame,
)

from database.database import (
    get_database,
    formatear_precio_inventario,
    normalizar_moneda,
)
from modules.kardex_movimiento_form import KardexMovimientoForm


class Kardex(QWidget):

    def __init__(self):
        super().__init__()
        self.db = get_database()
        self._inventario_id = None

        layout = QVBoxLayout(self)

        titulo = QLabel("KARDEX DE INVENTARIO (PEPS / FIFO)")
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("""
            font-size:26px;
            font-weight:bold;
            color:#2563EB;
            margin:10px;
        """)
        layout.addWidget(titulo)

        filtro = QHBoxLayout()
        filtro.addWidget(QLabel("Repuesto:"))
        self.combo_repuesto = QComboBox()
        self.combo_repuesto.setMinimumWidth(420)
        filtro.addWidget(self.combo_repuesto, 1)
        layout.addLayout(filtro)

        self.panel_kpi = QFrame()
        self.panel_kpi.setStyleSheet("""
            QFrame {
                background:#E2E8F0;
                border-radius:8px;
                padding:8px;
            }
            QLabel#kpiValor {
                font-size:18px;
                font-weight:bold;
                color:#1E293B;
            }
            QLabel#kpiTitulo {
                font-size:11px;
                color:#475569;
            }
        """)
        kpi_layout = QHBoxLayout(self.panel_kpi)

        self.lbl_stock_t = QLabel("Stock actual")
        self.lbl_stock_t.setObjectName("kpiTitulo")
        self.lbl_stock = QLabel("—")
        self.lbl_stock.setObjectName("kpiValor")

        self.lbl_prom_t = QLabel("Costo promedio")
        self.lbl_prom_t.setObjectName("kpiTitulo")
        self.lbl_prom = QLabel("—")
        self.lbl_prom.setObjectName("kpiValor")

        self.lbl_saldo_t = QLabel("Saldo (últ. mov.)")
        self.lbl_saldo_t.setObjectName("kpiTitulo")
        self.lbl_saldo = QLabel("—")
        self.lbl_saldo.setObjectName("kpiValor")

        for tit, val in (
            (self.lbl_stock_t, self.lbl_stock),
            (self.lbl_prom_t, self.lbl_prom),
            (self.lbl_saldo_t, self.lbl_saldo),
        ):
            col = QVBoxLayout()
            col.addWidget(tit)
            col.addWidget(val)
            kpi_layout.addLayout(col)

        layout.addWidget(self.panel_kpi)

        self.tabla = QTableWidget()
        self.tabla.setColumnCount(9)
        self.tabla.setHorizontalHeaderLabels([
            "Fecha",
            "Tipo",
            "Cantidad",
            "Costo unit.",
            "Costo total",
            "Saldo",
            "Costo prom.",
            "Documento",
            "Observación",
        ])
        self.tabla.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.tabla)

        botones = QHBoxLayout()
        self.btn_entrada = QPushButton("➕ Entrada")
        self.btn_salida = QPushButton("➖ Salida")
        self.btn_excel = QPushButton("📄 Excel")
        self.btn_pdf = QPushButton("🖨 PDF")

        botones.addWidget(self.btn_entrada)
        botones.addWidget(self.btn_salida)
        botones.addStretch()
        botones.addWidget(self.btn_excel)
        botones.addWidget(self.btn_pdf)
        layout.addLayout(botones)

        self.btn_entrada.clicked.connect(lambda: self._movimiento("ENTRADA"))
        self.btn_salida.clicked.connect(lambda: self._movimiento("SALIDA"))
        self.btn_excel.clicked.connect(self.exportar_excel)
        self.btn_pdf.clicked.connect(self.exportar_pdf)
        self.combo_repuesto.currentIndexChanged.connect(self._on_repuesto_cambiado)

        self._cargar_combo_repuestos()

    def _cargar_combo_repuestos(self):
        self.combo_repuesto.blockSignals(True)
        self.combo_repuesto.clear()
        self.combo_repuesto.addItem("— Seleccione repuesto —", None)
        for rep in self.db.obtener_repuestos():
            texto = f"{rep['codigo']} — {rep['descripcion']}"
            self.combo_repuesto.addItem(texto, rep["id"])
        self.combo_repuesto.blockSignals(False)
        self._on_repuesto_cambiado()

    def _on_repuesto_cambiado(self):
        self._inventario_id = self.combo_repuesto.currentData()
        habilitado = self._inventario_id is not None
        self.btn_entrada.setEnabled(habilitado)
        self.btn_salida.setEnabled(habilitado)
        self.btn_excel.setEnabled(habilitado)
        self.btn_pdf.setEnabled(habilitado)
        self.actualizar_vista()

    def actualizar_vista(self):
        if not self._inventario_id:
            self.lbl_stock.setText("—")
            self.lbl_prom.setText("—")
            self.lbl_saldo.setText("—")
            self.tabla.setRowCount(0)
            return

        resumen = self.db.obtener_resumen_kardex(self._inventario_id)
        moneda = resumen["moneda"]
        self.lbl_stock.setText(f"{resumen['stock_actual']:.2f}")
        self.lbl_prom.setText(
            formatear_precio_inventario(resumen["costo_promedio"], moneda)
        )
        self.lbl_saldo.setText(f"{resumen['saldo']:.2f}")

        movs = self.db.obtener_movimientos_kardex(self._inventario_id)
        self.tabla.setRowCount(len(movs))

        for i, mov in enumerate(movs):
            m = normalizar_moneda(
                mov["moneda"] if "moneda" in mov.keys() else moneda
            )
            fila = [
                mov["fecha"],
                mov["tipo"],
                f"{float(mov['cantidad']):.2f}",
                formatear_precio_inventario(mov["costo_unitario"], m),
                formatear_precio_inventario(mov["costo_total"], m),
                f"{float(mov['saldo_cantidad']):.2f}",
                formatear_precio_inventario(mov["costo_promedio"], m),
                mov["documento"] or "",
                mov["observacion"] or "",
            ]
            for j, texto in enumerate(fila):
                self.tabla.setItem(i, j, QTableWidgetItem(str(texto)))

    def _movimiento(self, tipo: str):
        if not self._inventario_id:
            QMessageBox.warning(
                self,
                "Kardex",
                "Seleccione un repuesto.",
            )
            return
        dlg = KardexMovimientoForm(
            self,
            inventario_id=self._inventario_id,
            tipo=tipo,
        )
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.actualizar_vista()

    def exportar_excel(self):
        if not self._inventario_id:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            QMessageBox.critical(
                self,
                "Excel",
                "Instale openpyxl: pip install openpyxl",
            )
            return

        resumen = self.db.obtener_resumen_kardex(self._inventario_id)
        movs = self.db.obtener_movimientos_kardex(self._inventario_id)
        if not movs:
            QMessageBox.information(self, "Excel", "No hay movimientos.")
            return

        cod = resumen.get("codigo", "repuesto")
        nombre = f"kardex_{cod}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar Kardex",
            str(Path.home() / nombre),
            "Excel (*.xlsx)",
        )
        if not ruta:
            return
        if not ruta.lower().endswith(".xlsx"):
            ruta += ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Kardex"
        ws.append(["Repuesto", f"{resumen.get('codigo')} — {resumen.get('descripcion')}"])
        ws.append(["Stock actual", resumen["stock_actual"]])
        ws.append(["Costo promedio", formatear_precio_inventario(
            resumen["costo_promedio"], resumen["moneda"])])
        ws.append(["Saldo", resumen["saldo"]])
        ws.append([])

        headers = [
            "Fecha", "Tipo", "Cantidad", "Costo unit.", "Costo total",
            "Saldo", "Costo prom.", "Documento", "Observación",
        ]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)

        moneda = resumen["moneda"]
        for mov in movs:
            m = mov["moneda"] if "moneda" in mov.keys() else moneda
            ws.append([
                mov["fecha"],
                mov["tipo"],
                float(mov["cantidad"]),
                formatear_precio_inventario(mov["costo_unitario"], m),
                formatear_precio_inventario(mov["costo_total"], m),
                float(mov["saldo_cantidad"]),
                formatear_precio_inventario(mov["costo_promedio"], m),
                mov["documento"] or "",
                mov["observacion"] or "",
            ])

        wb.save(ruta)
        QMessageBox.information(self, "Excel", f"Exportado:\n{ruta}")

    def exportar_pdf(self):
        if not self._inventario_id:
            return

        resumen = self.db.obtener_resumen_kardex(self._inventario_id)
        movs = self.db.obtener_movimientos_kardex(self._inventario_id)

        cod = resumen.get("codigo", "repuesto")
        nombre = f"kardex_{cod}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar Kardex PDF",
            str(Path.home() / nombre),
            "PDF (*.pdf)",
        )
        if not ruta:
            return
        if not ruta.lower().endswith(".pdf"):
            ruta += ".pdf"

        moneda = resumen["moneda"]
        filas_html = ""
        for mov in movs:
            m = mov["moneda"] if "moneda" in mov.keys() else moneda
            filas_html += f"""
            <tr>
              <td>{mov['fecha']}</td>
              <td>{mov['tipo']}</td>
              <td align="right">{float(mov['cantidad']):.2f}</td>
              <td align="right">{formatear_precio_inventario(mov['costo_unitario'], m)}</td>
              <td align="right">{formatear_precio_inventario(mov['costo_total'], m)}</td>
              <td align="right">{float(mov['saldo_cantidad']):.2f}</td>
              <td align="right">{formatear_precio_inventario(mov['costo_promedio'], m)}</td>
              <td>{mov['documento'] or ''}</td>
            </tr>
            """

        html = f"""
        <html><head><meta charset="utf-8"/>
        <style>
          body {{ font-family: Segoe UI, Arial; font-size: 10pt; }}
          h1 {{ color: #2563EB; font-size: 16pt; }}
          table {{ border-collapse: collapse; width: 100%; }}
          th, td {{ border: 1px solid #ccc; padding: 4px; }}
          th {{ background: #E2E8F0; }}
        </style></head><body>
        <h1>Kardex — SGM Piedra Vial</h1>
        <p><b>Repuesto:</b> {resumen.get('codigo')} — {resumen.get('descripcion')}</p>
        <p>
          <b>Stock actual:</b> {resumen['stock_actual']:.2f} &nbsp;|&nbsp;
          <b>Costo promedio:</b> {formatear_precio_inventario(resumen['costo_promedio'], moneda)} &nbsp;|&nbsp;
          <b>Saldo:</b> {resumen['saldo']:.2f}
        </p>
        <table>
          <tr>
            <th>Fecha</th><th>Tipo</th><th>Cant.</th><th>C. unit.</th>
            <th>C. total</th><th>Saldo</th><th>C. prom.</th><th>Documento</th>
          </tr>
          {filas_html if filas_html else '<tr><td colspan="8">Sin movimientos</td></tr>'}
        </table>
        </body></html>
        """

        doc = QTextDocument()
        doc.setHtml(html)
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(ruta)
        doc.print_(printer)

        QMessageBox.information(self, "PDF", f"Exportado:\n{ruta}")
