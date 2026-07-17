from datetime import datetime
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QLineEdit,
    QMessageBox,
    QFileDialog,
    QDialog,
)

from database.database import get_database, formatear_precio_inventario
from modules.modules.inventario_form import InventarioForm


class Inventario(QWidget):

    ROLE_ID = Qt.ItemDataRole.UserRole

    def __init__(self):
        super().__init__()

        self.db = get_database()

        layout = QVBoxLayout(self)

        titulo = QLabel("INVENTARIO DE REPUESTOS")
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("""
            font-size:26px;
            font-weight:bold;
            color:#2563EB;
            margin:10px;
        """)
        layout.addWidget(titulo)

        busqueda_layout = QHBoxLayout()
        busqueda_layout.addWidget(QLabel("Buscar:"))
        self.campo_busqueda = QLineEdit()
        self.campo_busqueda.setPlaceholderText(
            "Código, descripción, marca, equipo, proveedor..."
        )
        self.campo_busqueda.setClearButtonEnabled(True)
        busqueda_layout.addWidget(self.campo_busqueda)
        layout.addLayout(busqueda_layout)

        self.tabla = QTableWidget()
        self.tabla.setColumnCount(12)
        self.tabla.setHorizontalHeaderLabels([
            "Código",
            "Descripción",
            "N° Parte",
            "Marca",
            "Equipo",
            "Categoría",
            "Ubicación",
            "Stock",
            "Stock Min.",
            "Unidad",
            "Precio",
            "Proveedor",
        ])
        self.tabla.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )
        self.tabla.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )
        layout.addWidget(self.tabla)

        botones = QHBoxLayout()
        self.btnNuevo = QPushButton("➕ Nuevo")
        self.btnEditar = QPushButton("✏ Editar")
        self.btnEliminar = QPushButton("🗑 Eliminar")
        self.btnBuscar = QPushButton("🔎 Buscar")
        self.btnExcel = QPushButton("📄 Exportar Excel")

        botones.addWidget(self.btnNuevo)
        botones.addWidget(self.btnEditar)
        botones.addWidget(self.btnEliminar)
        botones.addWidget(self.btnBuscar)
        botones.addStretch()
        botones.addWidget(self.btnExcel)
        layout.addLayout(botones)

        self.btnNuevo.clicked.connect(self.nuevo)
        self.btnEditar.clicked.connect(self.editar)
        self.btnEliminar.clicked.connect(self.eliminar)
        self.btnBuscar.clicked.connect(self.enfocar_busqueda)
        self.btnExcel.clicked.connect(self.exportar_excel)
        self.campo_busqueda.textChanged.connect(self.buscar_en_tiempo_real)

        self.cargar_datos()

    def enfocar_busqueda(self):
        self.campo_busqueda.setFocus()
        self.campo_busqueda.selectAll()

    def _filas_desde_bd(self):
        termino = self.campo_busqueda.text()
        return self.db.buscar_repuestos(termino)

    def cargar_datos(self):
        filas = self._filas_desde_bd()
        self.tabla.setRowCount(len(filas))

        for fila_idx, row in enumerate(filas):
            moneda = row["moneda"] if "moneda" in row.keys() else "USD"
            precio_texto = formatear_precio_inventario(row["precio"], moneda)
            valores = [
                row["codigo"],
                row["descripcion"],
                row["numero_parte"],
                row["marca"],
                row["equipo"],
                row["categoria"],
                row["ubicacion"],
                row["stock"],
                row["stock_min"],
                row["unidad"],
                precio_texto,
                row["proveedor"],
            ]
            for col_idx, valor in enumerate(valores):
                texto = "" if valor is None else str(valor)
                item = QTableWidgetItem(texto)
                if col_idx == 0:
                    item.setData(self.ROLE_ID, row["id"])
                self.tabla.setItem(fila_idx, col_idx, item)

    def buscar_en_tiempo_real(self):
        self.cargar_datos()

    def _repuesto_id_seleccionado(self):
        filas = self.tabla.selectionModel().selectedRows()
        if not filas:
            return None
        fila = filas[0].row()
        item = self.tabla.item(fila, 0)
        if item is None:
            return None
        return item.data(self.ROLE_ID)

    def nuevo(self):
        dialogo = InventarioForm(self)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self.cargar_datos()

    def editar(self):
        repuesto_id = self._repuesto_id_seleccionado()
        if repuesto_id is None:
            QMessageBox.warning(
                self,
                "Editar",
                "Seleccione un repuesto de la tabla.",
            )
            return

        dialogo = InventarioForm(self, repuesto_id=repuesto_id)
        if dialogo.exec() == QDialog.DialogCode.Accepted:
            self.cargar_datos()

    def eliminar(self):
        repuesto_id = self._repuesto_id_seleccionado()
        if repuesto_id is None:
            QMessageBox.warning(
                self,
                "Eliminar",
                "Seleccione un repuesto de la tabla.",
            )
            return

        fila = self.tabla.selectionModel().selectedRows()[0].row()
        codigo = self.tabla.item(fila, 0).text()
        descripcion = self.tabla.item(fila, 1).text()

        confirmar = QMessageBox.question(
            self,
            "Confirmar eliminación",
            f"¿Eliminar el repuesto?\n\n{codigo} — {descripcion}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if confirmar != QMessageBox.StandardButton.Yes:
            return

        try:
            self.db.eliminar_repuesto(repuesto_id)
            self.cargar_datos()
            QMessageBox.information(
                self,
                "Eliminado",
                "Repuesto eliminado correctamente.",
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def exportar_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            QMessageBox.critical(
                self,
                "Exportar Excel",
                "Instale openpyxl: pip install openpyxl",
            )
            return

        filas = self.db.obtener_repuestos()
        if not filas:
            QMessageBox.information(
                self,
                "Exportar Excel",
                "No hay repuestos para exportar.",
            )
            return

        nombre_default = (
            f"inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar inventario a Excel",
            str(Path.home() / nombre_default),
            "Excel (*.xlsx)",
        )
        if not ruta:
            return
        if not ruta.lower().endswith(".xlsx"):
            ruta += ".xlsx"

        encabezados = [
            "Código",
            "Descripción",
            "N° Parte",
            "Marca",
            "Equipo",
            "Categoría",
            "Ubicación",
            "Stock",
            "Stock Min.",
            "Unidad",
            "Precio",
            "Proveedor",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        for col, titulo in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col, value=titulo)
            cell.font = Font(bold=True)

        for row_idx, row in enumerate(filas, start=2):
            ws.cell(row=row_idx, column=1, value=row["codigo"])
            ws.cell(row=row_idx, column=2, value=row["descripcion"])
            ws.cell(row=row_idx, column=3, value=row["numero_parte"])
            ws.cell(row=row_idx, column=4, value=row["marca"])
            ws.cell(row=row_idx, column=5, value=row["equipo"])
            ws.cell(row=row_idx, column=6, value=row["categoria"])
            ws.cell(row=row_idx, column=7, value=row["ubicacion"])
            ws.cell(row=row_idx, column=8, value=row["stock"])
            ws.cell(row=row_idx, column=9, value=row["stock_min"])
            ws.cell(row=row_idx, column=10, value=row["unidad"])
            moneda = row["moneda"] if "moneda" in row.keys() else "USD"
            ws.cell(
                row=row_idx,
                column=11,
                value=formatear_precio_inventario(row["precio"], moneda),
            )
            ws.cell(row=row_idx, column=12, value=row["proveedor"])

        wb.save(ruta)
        QMessageBox.information(
            self,
            "Exportar Excel",
            f"Inventario exportado correctamente:\n{ruta}",
        )
